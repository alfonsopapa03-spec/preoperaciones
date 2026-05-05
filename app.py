import streamlit as st
import psycopg2
import pandas as pd
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pytz

# ==================== CONFIGURACIÓN ====================
st.set_page_config(
    page_title="Inspección Preoperacional",
    layout="wide",
    page_icon="🔧",
    initial_sidebar_state="collapsed"
)

# ==================== CREDENCIALES ====================
SUPABASE_DB_URL = "postgresql://postgres.ogfenizdijcboekqhuhd:Conejito200$@aws-1-us-west-2.pooler.supabase.com:6543/postgres"

# ==================== CATÁLOGO DE MÁQUINAS ====================
# ⚙️ CAMBIA AQUÍ LOS NOMBRES REALES DE TUS 11 MÁQUINAS
MAQUINAS = [
    "Máquina 01",
    "Máquina 02",
    "Máquina 03",
    "Máquina 04",
    "Máquina 05",
    "Máquina 06",
    "Máquina 07",
    "Máquina 08",
    "Máquina 09",
    "Máquina 10",
    "Máquina 11",
]

# ==================== OPERADORES ====================
# ⚙️ CAMBIA AQUÍ LOS NOMBRES DE TUS OPERADORES
OPERADORES = sorted([
    "Operador 01",
    "Operador 02",
    "Operador 03",
    "Operador 04",
    "Operador 05",
])

REVISORES = sorted([
    "Supervisor 01",
    "Supervisor 02",
    "Supervisor 03",
])

# ==================== CHECKLIST — ITEMS POR SECCIÓN ====================
CHECKLIST = {
    "ANTES DE SU USO": [
        "¿Ha sido capacitado el trabajador para utilizar la máquina?",
        "¿Tiene permiso el trabajador para utilizar la máquina?",
        "¿Se ha verificado que la presión del aire se encuentre en 125 PSI?",
        "¿Se ha verificado que los desenrolladores contengan material?",
        "¿Se ha inspeccionado que los electro/válvulas funcionen adecuadamente?",
        "¿Se ha comprobado que los ganchos de ajuste funcionen?",
        "¿El 'carro' de tracción del material funciona satisfactoriamente?",
        "¿Se ha verificado que el tape se encuentre en óptimas condiciones?",
        "¿Se ha inspeccionado que las palancas de emergencias (6 en total) funcionen correctamente?",
        "¿Se ha verificado el estado de los cabezales (inferior/superior)?",
        "¿El nivel de lubricación se encontrará en nivel correspondiente?",
        "¿El manómetro de aire se encuentra funcionando correctamente?",
        "¿El nivel de aceite del líquido refrigerante se encontrará en nivel adecuado?",
        "¿Se ha ajustado la altura del panel a medida correspondiente?",
    ],
    "INSPECCIÓN DEL LUGAR DE TRABAJO": [
        "¿Se ha inspeccionado el lugar de trabajo? (material combustible, riesgo de incendios, instalaciones, otros trabajadores, etc.)",
        "¿La iluminación del área de trabajo es adecuada para operación de la máquina sin riesgos?",
        "¿Se ha inspeccionado que el área esté limpia y libre de obstáculos?",
    ],
    "ELEMENTOS DE PROTECCIÓN PERSONAL (EPP)": [
        "¿Cuenta con los elementos de protección personal? (protector de ojos, oídos, guantes y calzado)",
        "¿El trabajador está vestido apropiadamente? (Camisa manga larga, pantalón de dotación y calzado de seguridad)",
        "¿Se evidencia el NO uso de joyas, relojes y ropa holgada?",
    ],
    "SEGURIDAD ELÉCTRICA": [
        "¿Se tiene el cabello recogido si lo tiene largo?",
        "¿Se ha verificado que el cable de alimentación esté en buen estado?",
        "¿Se ha revisado que el enchufe se encuentre en buenas condiciones?",
        "¿El interruptor de encendido funciona correctamente?",
    ],
}

TODAS_PREGUNTAS = []
for seccion, preguntas in CHECKLIST.items():
    for p in preguntas:
        TODAS_PREGUNTAS.append((seccion, p))

# ==================== CSS ====================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Rajdhani:wght@400;500;600;700&family=Noto+Sans:wght@300;400;500&display=swap');

    html, body, [class*="css"] {
        font-family: 'Noto Sans', sans-serif;
    }

    .main-header {
        background: linear-gradient(135deg, #0a0a0a 0%, #1a1a2e 50%, #16213e 100%);
        padding: 1.8rem 2.5rem;
        border-radius: 14px;
        margin-bottom: 1.5rem;
        border: 1px solid #0f3460;
        position: relative;
        overflow: hidden;
    }
    .main-header::before {
        content: '';
        position: absolute;
        top: 0; left: 0; right: 0; bottom: 0;
        background: repeating-linear-gradient(
            45deg,
            transparent,
            transparent 10px,
            rgba(255,165,0,0.02) 10px,
            rgba(255,165,0,0.02) 11px
        );
    }
    .main-header h1 {
        font-family: 'Rajdhani', sans-serif;
        font-size: 2.2rem;
        font-weight: 700;
        color: #ffffff;
        margin: 0;
        letter-spacing: 3px;
        text-transform: uppercase;
    }
    .main-header .accent { color: #f5a623; }
    .main-header p { color: #8899aa; margin: 0.3rem 0 0; font-size: 0.85rem; letter-spacing: 1px; }

    .seccion-header {
        background: linear-gradient(90deg, #0f3460, #16213e);
        color: #f5a623;
        font-family: 'Rajdhani', sans-serif;
        font-size: 1rem;
        font-weight: 700;
        letter-spacing: 2px;
        text-transform: uppercase;
        padding: 0.6rem 1.2rem;
        border-radius: 6px;
        border-left: 4px solid #f5a623;
        margin: 1rem 0 0.5rem;
    }

    .item-pregunta {
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        border-radius: 8px;
        padding: 0.7rem 1rem;
        margin-bottom: 0.4rem;
        font-size: 0.88rem;
        color: #2d3748;
        line-height: 1.4;
    }

    .badge-cumple {
        background: #d4edda; color: #155724;
        padding: 2px 8px; border-radius: 12px;
        font-size: 0.75rem; font-weight: 600;
    }
    .badge-nocumple {
        background: #f8d7da; color: #721c24;
        padding: 2px 8px; border-radius: 12px;
        font-size: 0.75rem; font-weight: 600;
    }
    .badge-na {
        background: #e2e3e5; color: #383d41;
        padding: 2px 8px; border-radius: 12px;
        font-size: 0.75rem; font-weight: 600;
    }

    .kpi-card {
        background: white;
        border-radius: 10px;
        padding: 1rem 1.2rem;
        border-left: 5px solid #f5a623;
        box-shadow: 0 2px 8px rgba(0,0,0,0.07);
        margin-bottom: 0.5rem;
    }
    .kpi-val { font-size: 2rem; font-weight: 700; color: #0a0a0a; font-family: 'Rajdhani', sans-serif; }
    .kpi-lbl { font-size: 0.75rem; color: #666; text-transform: uppercase; letter-spacing: 1px; }

    .alerta-nc {
        background: #fff5f5;
        border: 1px solid #fc8181;
        border-left: 4px solid #e53e3e;
        border-radius: 8px;
        padding: 0.8rem 1rem;
        margin: 0.5rem 0;
        font-size: 0.85rem;
        color: #742a2a;
    }

    div[data-testid="stTabs"] button {
        font-family: 'Rajdhani', sans-serif;
        font-weight: 600;
        font-size: 1rem;
        letter-spacing: 1px;
    }

    .resultado-cumple {
        background: linear-gradient(135deg, #d4edda, #c3e6cb);
        border: 1px solid #28a745;
        border-radius: 10px;
        padding: 1rem;
        text-align: center;
        font-family: 'Rajdhani', sans-serif;
        font-size: 1.5rem;
        color: #155724;
        font-weight: 700;
    }
    .resultado-nc {
        background: linear-gradient(135deg, #f8d7da, #f5c6cb);
        border: 1px solid #dc3545;
        border-radius: 10px;
        padding: 1rem;
        text-align: center;
        font-family: 'Rajdhani', sans-serif;
        font-size: 1.5rem;
        color: #721c24;
        font-weight: 700;
    }
</style>
""", unsafe_allow_html=True)

# ==================== BASE DE DATOS ====================
class DB:
    def __init__(self):
        self.url = SUPABASE_DB_URL
        self.init()

    def conn(self):
        return psycopg2.connect(self.url)

    def init(self):
        try:
            c = self.conn()
            cur = c.cursor()
            # Tabla principal de inspecciones
            cur.execute("""
                CREATE TABLE IF NOT EXISTS inspecciones_preoperacionales (
                    id SERIAL PRIMARY KEY,
                    fecha_registro TIMESTAMP DEFAULT (now() AT TIME ZONE 'America/Bogota'),
                    fecha DATE NOT NULL,
                    maquina TEXT NOT NULL,
                    modelo TEXT,
                    marca TEXT,
                    placa TEXT,
                    operador TEXT,
                    revisor TEXT,
                    cliente_proyecto TEXT,
                    responsable_mantenimiento TEXT,
                    observaciones TEXT,
                    total_items INTEGER DEFAULT 0,
                    items_cumple INTEGER DEFAULT 0,
                    items_no_cumple INTEGER DEFAULT 0,
                    items_na INTEGER DEFAULT 0,
                    porcentaje_cumplimiento NUMERIC(5,2) DEFAULT 0
                )
            """)
            # Tabla de respuestas por ítem
            cur.execute("""
                CREATE TABLE IF NOT EXISTS inspecciones_items (
                    id SERIAL PRIMARY KEY,
                    inspeccion_id INTEGER REFERENCES inspecciones_preoperacionales(id) ON DELETE CASCADE,
                    seccion TEXT NOT NULL,
                    pregunta TEXT NOT NULL,
                    respuesta TEXT NOT NULL,
                    observacion_item TEXT
                )
            """)
            c.commit()
            c.close()
        except Exception as e:
            st.error(f"Error DB init: {e}")

    def guardar_inspeccion(self, datos: dict, items: list) -> bool:
        try:
            c = self.conn()
            cur = c.cursor()
            cur.execute("""
                INSERT INTO inspecciones_preoperacionales
                (fecha, maquina, modelo, marca, placa, operador, revisor,
                 cliente_proyecto, responsable_mantenimiento, observaciones,
                 total_items, items_cumple, items_no_cumple, items_na, porcentaje_cumplimiento)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id
            """, (
                datos["fecha"], datos["maquina"], datos["modelo"], datos["marca"],
                datos["placa"], datos["operador"], datos["revisor"],
                datos["cliente_proyecto"], datos["responsable_mantenimiento"],
                datos["observaciones"],
                datos["total_items"], datos["items_cumple"],
                datos["items_no_cumple"], datos["items_na"],
                datos["porcentaje_cumplimiento"]
            ))
            inspeccion_id = cur.fetchone()[0]

            for item in items:
                cur.execute("""
                    INSERT INTO inspecciones_items (inspeccion_id, seccion, pregunta, respuesta, observacion_item)
                    VALUES (%s, %s, %s, %s, %s)
                """, (inspeccion_id, item["seccion"], item["pregunta"],
                      item["respuesta"], item.get("observacion_item", "")))

            c.commit()
            c.close()
            return True
        except Exception as e:
            st.error(f"Error guardando inspección: {e}")
            return False

    def obtener_inspecciones(self, fecha_ini=None, fecha_fin=None,
                              maquina=None, operador=None) -> pd.DataFrame:
        c = self.conn()
        q = """SELECT id, fecha, maquina, modelo, marca, placa, operador, revisor,
                      cliente_proyecto, responsable_mantenimiento,
                      total_items, items_cumple, items_no_cumple, items_na,
                      porcentaje_cumplimiento, observaciones
               FROM inspecciones_preoperacionales WHERE 1=1"""
        params = []
        if fecha_ini: q += " AND fecha >= %s"; params.append(fecha_ini)
        if fecha_fin: q += " AND fecha <= %s"; params.append(fecha_fin)
        if maquina and maquina != "Todas": q += " AND maquina = %s"; params.append(maquina)
        if operador: q += " AND operador ILIKE %s"; params.append(f"%{operador}%")
        q += " ORDER BY fecha DESC, id DESC"
        try:
            df = pd.read_sql(q, c, params=params)
            return df
        except:
            return pd.DataFrame()
        finally:
            c.close()

    def obtener_items_inspeccion(self, inspeccion_id: int) -> pd.DataFrame:
        c = self.conn()
        try:
            df = pd.read_sql(
                "SELECT seccion, pregunta, respuesta, observacion_item FROM inspecciones_items WHERE inspeccion_id=%s ORDER BY id",
                c, params=[inspeccion_id]
            )
            return df
        except:
            return pd.DataFrame()
        finally:
            c.close()

    def eliminar_inspeccion(self, inspeccion_id: int) -> bool:
        try:
            c = self.conn()
            cur = c.cursor()
            cur.execute("DELETE FROM inspecciones_preoperacionales WHERE id=%s", (inspeccion_id,))
            c.commit()
            c.close()
            return True
        except Exception as e:
            st.error(f"Error eliminando: {e}")
            return False

    def ya_existe_hoy(self, fecha, maquina) -> bool:
        c = self.conn()
        try:
            df = pd.read_sql(
                "SELECT id FROM inspecciones_preoperacionales WHERE fecha=%s AND maquina=%s",
                c, params=[fecha, maquina]
            )
            return len(df) > 0
        except:
            return False
        finally:
            c.close()

    def stats_dashboard(self, fecha_ini, fecha_fin):
        c = self.conn()
        try:
            df = pd.read_sql("""
                SELECT fecha, maquina, operador, items_cumple, items_no_cumple,
                       items_na, total_items, porcentaje_cumplimiento
                FROM inspecciones_preoperacionales
                WHERE fecha >= %s AND fecha <= %s
                ORDER BY fecha
            """, c, params=[fecha_ini, fecha_fin])
            return df
        except:
            return pd.DataFrame()
        finally:
            c.close()

    def items_nc_frecuentes(self, fecha_ini, fecha_fin):
        c = self.conn()
        try:
            df = pd.read_sql("""
                SELECT ii.seccion, ii.pregunta, COUNT(*) as veces
                FROM inspecciones_items ii
                JOIN inspecciones_preoperacionales ip ON ii.inspeccion_id = ip.id
                WHERE ip.fecha >= %s AND ip.fecha <= %s
                  AND ii.respuesta = 'NC'
                GROUP BY ii.seccion, ii.pregunta
                ORDER BY veces DESC
                LIMIT 15
            """, c, params=[fecha_ini, fecha_fin])
            return df
        except:
            return pd.DataFrame()
        finally:
            c.close()


# ==================== EXCEL ====================
def generar_excel(df: pd.DataFrame, db: 'DB', titulo: str = "Inspecciones Preoperacionales") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inspecciones"

    ft_titulo  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ft_header  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    ft_normal  = Font(name="Calibri", size=9)
    ft_total   = Font(name="Calibri", bold=True, size=10)
    ft_nc      = Font(name="Calibri", size=9, color="C0392B", bold=True)
    ft_cumple  = Font(name="Calibri", size=9, color="1A5C2A", bold=True)

    fill_titulo  = PatternFill("solid", start_color="0F2027")
    fill_header  = PatternFill("solid", start_color="203A43")
    fill_alt     = PatternFill("solid", start_color="EBF5FB")
    fill_total   = PatternFill("solid", start_color="D5DBDB")
    fill_nc_row  = PatternFill("solid", start_color="FADBD8")
    fill_ok_row  = PatternFill("solid", start_color="D5F5E3")

    borde  = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    izq    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    now_col = datetime.now(pytz.timezone("America/Bogota"))

    # ── HOJA 1: Resumen de inspecciones ──
    total_cols = 15
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws["A1"] = f"🔧 {titulo}   |   Generado: {now_col.strftime('%d/%m/%Y %H:%M')} (COL)   |   Total: {len(df)} inspecciones"
    ws["A1"].font = ft_titulo
    ws["A1"].fill = fill_titulo
    ws["A1"].alignment = centro
    ws.row_dimensions[1].height = 28

    columnas = [
        ("id",                       "ID",            6),
        ("fecha",                    "FECHA",         12),
        ("maquina",                  "MÁQUINA",       20),
        ("modelo",                   "MODELO",        14),
        ("marca",                    "MARCA",         14),
        ("placa",                    "PLACA",         12),
        ("operador",                 "OPERADOR",      24),
        ("revisor",                  "REVISOR",       22),
        ("cliente_proyecto",         "CLIENTE/PROYECTO", 22),
        ("total_items",              "TOTAL ÍTEMS",   10),
        ("items_cumple",             "✅ CUMPLE",     10),
        ("items_no_cumple",          "❌ NO CUMPLE",  11),
        ("items_na",                 "N/A",           8),
        ("porcentaje_cumplimiento",  "% CUMPLIMIENTO",14),
        ("observaciones",            "OBSERVACIONES", 30),
    ]

    for idx, (key, nombre, ancho) in enumerate(columnas, start=1):
        cell = ws.cell(row=2, column=idx, value=nombre)
        cell.font = ft_header
        cell.fill = fill_header
        cell.alignment = centro
        cell.border = borde
        ws.column_dimensions[get_column_letter(idx)].width = ancho
    ws.row_dimensions[2].height = 26

    for row_idx, (_, fila) in enumerate(df.iterrows(), start=3):
        pct = float(fila.get("porcentaje_cumplimiento", 0) or 0)
        es_nc = pct < 80
        fill_f = fill_nc_row if es_nc else (fill_alt if row_idx % 2 == 0 else None)

        for col_idx, (key, _, _) in enumerate(columnas, start=1):
            val = fila.get(key, "")
            if isinstance(val, float) and pd.isna(val): val = ""
            if key == "porcentaje_cumplimiento" and val != "":
                val = f"{float(val):.1f}%"
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val != "" else "")
            cell.border = borde
            cell.alignment = centro if key in ("id","fecha","total_items","items_cumple","items_no_cumple","items_na","porcentaje_cumplimiento","placa","modelo","marca") else izq
            if key == "porcentaje_cumplimiento":
                cell.font = ft_nc if es_nc else ft_cumple
            else:
                cell.font = ft_normal
            if fill_f: cell.fill = fill_f
        ws.row_dimensions[row_idx].height = 18

    total_row = len(df) + 3
    ws.merge_cells(f"A{total_row}:{get_column_letter(total_cols)}{total_row}")
    pct_prom = df["porcentaje_cumplimiento"].mean() if len(df) > 0 else 0
    ct = ws.cell(row=total_row, column=1,
                 value=f"TOTAL: {len(df)} inspecciones   |   Cumplimiento promedio: {pct_prom:.1f}%")
    ct.font = ft_total
    ct.fill = fill_total
    ct.alignment = centro
    ws.freeze_panes = "A3"

    # ── HOJA 2: Detalle por ítems ──
    ws2 = wb.create_sheet("Detalle Ítems")
    ws2["A1"] = "Detalle de Ítems por Inspección"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", start_color="0F2027")
    ws2["A1"].alignment = centro
    ws2.row_dimensions[1].height = 26

    hdrs2 = ["ID INSP.", "FECHA", "MÁQUINA", "OPERADOR", "SECCIÓN", "PREGUNTA", "RESPUESTA", "OBSERVACIÓN ÍTEM"]
    anchos2 = [8, 12, 20, 22, 28, 60, 12, 30]
    for ci, (h, w) in enumerate(zip(hdrs2, anchos2), start=1):
        c = ws2.cell(2, ci, h)
        c.font = ft_header
        c.fill = PatternFill("solid", start_color="203A43")
        c.alignment = centro
        c.border = borde
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[2].height = 22

    fila_det = 3
    fill_seccion_colors = {
        "ANTES DE SU USO": PatternFill("solid", start_color="EBF5FB"),
        "INSPECCIÓN DEL LUGAR DE TRABAJO": PatternFill("solid", start_color="FEF9E7"),
        "ELEMENTOS DE PROTECCIÓN PERSONAL (EPP)": PatternFill("solid", start_color="F9EBEA"),
        "SEGURIDAD ELÉCTRICA": PatternFill("solid", start_color="EAFAF1"),
    }

    for _, fila_insp in df.iterrows():
        items_df = db.obtener_items_inspeccion(int(fila_insp["id"]))
        if items_df.empty:
            continue
        for _, item in items_df.iterrows():
            fill_c = fill_seccion_colors.get(str(item.get("seccion", "")))
            resp = str(item.get("respuesta", ""))
            if resp == "NC":
                fill_c = fill_nc_row

            vals = [
                str(fila_insp["id"]),
                str(fila_insp["fecha"]),
                str(fila_insp["maquina"]),
                str(fila_insp["operador"]),
                str(item.get("seccion", "")),
                str(item.get("pregunta", "")),
                resp,
                str(item.get("observacion_item", "") or ""),
            ]
            for ci, v in enumerate(vals, start=1):
                cell = ws2.cell(fila_det, ci, v)
                cell.border = borde
                cell.font = ft_nc if resp == "NC" else ft_normal
                cell.alignment = centro if ci in (1, 2, 7) else izq
                if fill_c: cell.fill = fill_c
            ws2.row_dimensions[fila_det].height = 16
            fila_det += 1

    ws2.freeze_panes = "A3"

    # ── HOJA 3: Resumen por Máquina ──
    ws3 = wb.create_sheet("Por Máquina")
    ws3["A1"] = "Resumen de Cumplimiento por Máquina"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", start_color="0F2027")
    ws3["A1"].alignment = centro
    ws3.row_dimensions[1].height = 26

    hdrs3 = ["MÁQUINA", "INSPECCIONES", "% PROM. CUMPL.", "MIN %", "MAX %", "NC TOTAL"]
    anchos3 = [24, 12, 16, 10, 10, 10]
    for ci, (h, w) in enumerate(zip(hdrs3, anchos3), start=1):
        c = ws3.cell(2, ci, h)
        c.font = ft_header
        c.fill = PatternFill("solid", start_color="203A43")
        c.alignment = centro
        c.border = borde
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.row_dimensions[2].height = 22

    if "maquina" in df.columns and not df.empty:
        df_maq = df.groupby("maquina").agg(
            inspecciones=("id", "count"),
            pct_prom=("porcentaje_cumplimiento", "mean"),
            pct_min=("porcentaje_cumplimiento", "min"),
            pct_max=("porcentaje_cumplimiento", "max"),
            nc_total=("items_no_cumple", "sum"),
        ).reset_index().sort_values("pct_prom")

        for i, row in enumerate(df_maq.itertuples(), start=3):
            fill_c = PatternFill("solid", start_color="FADBD8") if row.pct_prom < 80 else (
                PatternFill("solid", start_color="EBF5FB") if i % 2 == 0 else None)
            vals = [row.maquina, row.inspecciones,
                    f"{row.pct_prom:.1f}%", f"{row.pct_min:.1f}%",
                    f"{row.pct_max:.1f}%", int(row.nc_total)]
            for ci, v in enumerate(vals, start=1):
                cell = ws3.cell(i, ci, v)
                cell.font = ft_nc if (ci == 3 and row.pct_prom < 80) else ft_normal
                cell.border = borde
                cell.alignment = izq if ci == 1 else centro
                if fill_c: cell.fill = fill_c

    ws3.freeze_panes = "A3"

    # ── HOJA 4: Top NC ──
    ws4 = wb.create_sheet("Top No Cumple")
    ws4["A1"] = "Ítems con Más Incumplimientos"
    ws4["A1"].font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ws4["A1"].fill = PatternFill("solid", start_color="0F2027")
    ws4["A1"].alignment = centro
    ws4.row_dimensions[1].height = 26

    hdrs4 = ["#", "SECCIÓN", "PREGUNTA", "VECES NC"]
    anchos4 = [5, 28, 70, 10]
    for ci, (h, w) in enumerate(zip(hdrs4, anchos4), start=1):
        c = ws4.cell(2, ci, h)
        c.font = ft_header
        c.fill = PatternFill("solid", start_color="922B21")
        c.alignment = centro
        c.border = borde
        ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.row_dimensions[2].height = 22

    # Calcular Top NC desde items
    nc_data = []
    for _, fila_insp in df.iterrows():
        items_df = db.obtener_items_inspeccion(int(fila_insp["id"]))
        if not items_df.empty:
            nc = items_df[items_df["respuesta"] == "NC"]
            for _, r in nc.iterrows():
                nc_data.append({"seccion": r["seccion"], "pregunta": r["pregunta"]})

    if nc_data:
        df_nc = pd.DataFrame(nc_data)
        df_nc_agg = df_nc.groupby(["seccion", "pregunta"]).size().reset_index(name="veces").sort_values("veces", ascending=False).head(20)
        for i, row in enumerate(df_nc_agg.itertuples(), start=3):
            fill_c = PatternFill("solid", start_color="FADBD8") if i <= 4 else (
                PatternFill("solid", start_color="FEF9E7") if i % 2 == 0 else None)
            vals = [i - 2, row.seccion, row.pregunta, row.veces]
            for ci, v in enumerate(vals, start=1):
                cell = ws4.cell(i, ci, v)
                cell.font = ft_nc if i <= 4 else ft_normal
                cell.border = borde
                cell.alignment = centro if ci in (1, 4) else izq
                if fill_c: cell.fill = fill_c

    ws4.freeze_panes = "A3"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ==================== MAIN ====================
def main():
    st.markdown("""
    <div class="main-header">
        <h1>🔧 REGISTRO PRE<span class="accent">OPERACIONAL</span> DE EQUIPO</h1>
        <p>CONTROL DE INSPECCIONES · SCA ZF · TRAZABILIDAD TOTAL</p>
    </div>
    """, unsafe_allow_html=True)

    if "db" not in st.session_state:
        st.session_state.db = DB()
    if "viendo_id" not in st.session_state:
        st.session_state.viendo_id = None

    db = st.session_state.db

    tab1, tab2, tab3 = st.tabs(["📋 Nueva Inspección", "🔍 Historial y Reportes", "📊 Dashboard"])

    # ======================== TAB 1: NUEVA INSPECCIÓN ========================
    with tab1:
        st.markdown("### Registrar Inspección Preoperacional")

        # ── Datos del equipo ──
        st.markdown('<div class="seccion-header">1. DATOS DEL EQUIPO</div>', unsafe_allow_html=True)
        d1, d2, d3, d4 = st.columns(4)
        with d1:
            fecha = st.date_input("📅 Fecha", datetime.now(), key="n_fecha")
        with d2:
            maquina = st.selectbox("🏭 Máquina / Equipo", MAQUINAS, key="n_maq")
        with d3:
            modelo = st.text_input("Modelo", key="n_modelo", placeholder="Ej: ZF-2025")
        with d4:
            marca = st.text_input("Marca", key="n_marca", placeholder="Ej: SCA")

        d5, d6, d7, d8 = st.columns(4)
        with d5:
            placa = st.text_input("Placa / Serie", key="n_placa", placeholder="Ej: SCA-001")
        with d6:
            operador_sel = st.selectbox("👷 Operador", ["— Seleccionar —"] + OPERADORES, key="n_op")
        with d7:
            revisor_sel = st.selectbox("👁️ Revisado por", ["— Seleccionar —"] + REVISORES, key="n_rev")
        with d8:
            cliente_proy = st.text_input("Cliente / Proyecto", key="n_cli", placeholder="Nombre del cliente...")

        resp_mantenimiento = st.text_input("🔧 Responsable Mantenimiento", key="n_mant", placeholder="Nombre del responsable de mantenimiento...")

        # Verificar si ya existe inspección hoy
        if maquina:
            ya_existe = db.ya_existe_hoy(fecha, maquina)
            if ya_existe:
                st.warning(f"⚠️ Ya existe una inspección para **{maquina}** el **{fecha}**. Si continúas se creará un duplicado.")

        # ── Checklist ──
        st.markdown('<div class="seccion-header">2. LISTA DE ACTIVIDADES</div>', unsafe_allow_html=True)
        st.caption("Marca cada ítem: **C** = Cumple · **NC** = No Cumple · **N/A** = No Aplica")

        respuestas = {}
        obs_items = {}
        nc_detectados = []

        for seccion, preguntas in CHECKLIST.items():
            st.markdown(f'<div class="seccion-header" style="font-size:0.85rem; margin-top:0.8rem;">📌 {seccion}</div>', unsafe_allow_html=True)
            for i, pregunta in enumerate(preguntas):
                key_base = f"resp_{seccion[:10]}_{i}"
                col_preg, col_resp, col_obs = st.columns([5, 2, 3])
                with col_preg:
                    st.markdown(f'<div class="item-pregunta">{pregunta}</div>', unsafe_allow_html=True)
                with col_resp:
                    resp = st.radio(
                        "Respuesta",
                        options=["C", "NC", "N/A"],
                        index=0,
                        key=key_base,
                        horizontal=True,
                        label_visibility="collapsed"
                    )
                    respuestas[(seccion, pregunta)] = resp
                with col_obs:
                    obs_i = ""
                    if resp == "NC":
                        nc_detectados.append(pregunta[:60])
                        obs_i = st.text_input(
                            "Observación del ítem",
                            key=f"obs_{key_base}",
                            placeholder="Describir hallazgo...",
                            label_visibility="collapsed"
                        )
                    obs_items[(seccion, pregunta)] = obs_i

        # Resumen en tiempo real
        total = len(respuestas)
        cumple = sum(1 for v in respuestas.values() if v == "C")
        no_cumple = sum(1 for v in respuestas.values() if v == "NC")
        na = sum(1 for v in respuestas.values() if v == "N/A")
        efectivos = total - na
        pct = round(cumple / efectivos * 100, 1) if efectivos > 0 else 0

        st.divider()
        st.markdown("#### 📊 Resumen en tiempo real")
        r1, r2, r3, r4, r5 = st.columns(5)
        r1.metric("Total ítems", total)
        r2.metric("✅ Cumple", cumple)
        r3.metric("❌ No Cumple", no_cumple)
        r4.metric("➖ N/A", na)
        r5.metric("📈 % Cumplimiento", f"{pct}%")

        if no_cumple > 0:
            st.markdown(
                f'<div class="alerta-nc">⚠️ <strong>{no_cumple} ítem(s) con NO CUMPLE.</strong> '
                f'Reportar inmediatamente al encargado de equipos y al departamento de mantenimiento.<br>'
                f'<em>{" · ".join(nc_detectados[:5])}</em></div>',
                unsafe_allow_html=True
            )
        if pct == 100:
            st.markdown('<div class="resultado-cumple">✅ INSPECCIÓN 100% — EQUIPO APTO PARA OPERAR</div>', unsafe_allow_html=True)
        elif pct >= 80:
            st.markdown(f'<div class="resultado-cumple">✅ INSPECCIÓN APROBADA — {pct}% cumplimiento</div>', unsafe_allow_html=True)
        elif pct > 0:
            st.markdown(f'<div class="resultado-nc">❌ INSPECCIÓN CON OBSERVACIONES — {pct}% cumplimiento — REVISAR ANTES DE OPERAR</div>', unsafe_allow_html=True)

        # ── Sección 3: Datos de control ──
        st.markdown('<div class="seccion-header">3. OBSERVACIONES GENERALES Y DATOS DE CONTROL</div>', unsafe_allow_html=True)
        observaciones = st.text_area("📝 Comentarios / Observaciones generales", height=100, key="n_obs",
                                      placeholder="Cualquier anomalía, favor reportarla en el proyecto y al encargado de equipos...")

        st.divider()
        col_btn1, col_btn2 = st.columns([1, 3])
        with col_btn1:
            guardar = st.button("💾 Guardar Inspección", type="primary", use_container_width=True)

        if guardar:
            operador = "" if operador_sel == "— Seleccionar —" else operador_sel
            revisor  = "" if revisor_sel  == "— Seleccionar —" else revisor_sel

            if not operador:
                st.error("⚠️ Selecciona el operador antes de guardar.")
            else:
                datos = {
                    "fecha": fecha,
                    "maquina": maquina,
                    "modelo": modelo,
                    "marca": marca,
                    "placa": placa,
                    "operador": operador,
                    "revisor": revisor,
                    "cliente_proyecto": cliente_proy,
                    "responsable_mantenimiento": resp_mantenimiento,
                    "observaciones": observaciones,
                    "total_items": total,
                    "items_cumple": cumple,
                    "items_no_cumple": no_cumple,
                    "items_na": na,
                    "porcentaje_cumplimiento": pct,
                }
                items = [
                    {
                        "seccion": sec,
                        "pregunta": preg,
                        "respuesta": resp,
                        "observacion_item": obs_items.get((sec, preg), "")
                    }
                    for (sec, preg), resp in respuestas.items()
                ]
                if db.guardar_inspeccion(datos, items):
                    st.success(f"✅ Inspección guardada — {maquina} | {operador} | {fecha} | {pct}% cumplimiento")
                    st.balloons()

    # ======================== TAB 2: HISTORIAL ========================
    with tab2:
        st.markdown("### 🔍 Historial de Inspecciones")

        with st.expander("🛠️ Filtros", expanded=True):
            f1, f2, f3, f4 = st.columns(4)
            with f1: fi   = st.date_input("Desde", datetime.now() - timedelta(days=30), key="h_fi")
            with f2: ff   = st.date_input("Hasta", datetime.now(), key="h_ff")
            with f3:
                maq_f = st.selectbox("Máquina", ["Todas"] + MAQUINAS, key="h_maq")
            with f4:
                op_f = st.text_input("Operador", key="h_op")

        df = db.obtener_inspecciones(fi, ff, maq_f, op_f)

        if not df.empty:
            k1, k2, k3, k4 = st.columns(4)
            k1.metric("Total Inspecciones", len(df))
            k2.metric("✅ ≥ 80%", len(df[df["porcentaje_cumplimiento"] >= 80]))
            k3.metric("❌ < 80%", len(df[df["porcentaje_cumplimiento"] < 80]))
            k4.metric("📊 % Promedio", f"{df['porcentaje_cumplimiento'].mean():.1f}%")

            st.divider()

            col_e1, col_e2 = st.columns([2, 5])
            with col_e1:
                nombre_rep = st.text_input("Nombre del reporte", value="Inspecciones_Preoperacionales", key="rep_nom")
            with col_e2:
                st.markdown("<br>", unsafe_allow_html=True)
                excel_data = generar_excel(df, db, titulo=nombre_rep)
                st.download_button(
                    "⬇️ Descargar Excel",
                    data=excel_data,
                    file_name=f"{nombre_rep}_{datetime.now(pytz.timezone('America/Bogota')).strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )

            st.divider()

            cols_tabla = ["id", "fecha", "maquina", "operador", "revisor",
                          "items_cumple", "items_no_cumple", "porcentaje_cumplimiento"]
            cols_ex = [c for c in cols_tabla if c in df.columns]

            # Color condicional
            def color_pct(val):
                try:
                    v = float(str(val).replace("%",""))
                    color = "#d4edda" if v >= 80 else "#f8d7da"
                    return f"background-color: {color}"
                except:
                    return ""

            st.dataframe(
                df[cols_ex].style.applymap(color_pct, subset=["porcentaje_cumplimiento"]),
                use_container_width=True,
                hide_index=True
            )

            st.divider()
            st.subheader("🔎 Ver Detalle de Inspección")
            df["_label"] = df.apply(
                lambda r: f"ID {r['id']} | {r['fecha']} | {r['maquina']} | {r['operador']} | {r['porcentaje_cumplimiento']:.1f}%",
                axis=1
            )
            sel = st.selectbox("Seleccionar inspección:", df["_label"].tolist(), key="h_sel")

            if sel:
                vid = int(sel.split(" | ")[0].replace("ID ", ""))
                row = df[df["id"] == vid].iloc[0]

                c1, c2, c3 = st.columns(3)
                with c1:
                    st.info(f"**Máquina:** {row['maquina']}")
                    st.write(f"**Operador:** {row.get('operador','')}")
                    st.write(f"**Revisor:** {row.get('revisor','')}")
                    st.write(f"**Fecha:** {row['fecha']}")
                with c2:
                    st.write(f"**Modelo:** {row.get('modelo','')}")
                    st.write(f"**Marca:** {row.get('marca','')}")
                    st.write(f"**Placa:** {row.get('placa','')}")
                    st.write(f"**Cliente/Proyecto:** {row.get('cliente_proyecto','')}")
                with c3:
                    pct_v = float(row.get('porcentaje_cumplimiento', 0))
                    color = "🟢" if pct_v >= 80 else "🔴"
                    st.write(f"**{color} Cumplimiento:** {pct_v:.1f}%")
                    st.write(f"**✅ Cumple:** {row.get('items_cumple',0)}")
                    st.write(f"**❌ No Cumple:** {row.get('items_no_cumple',0)}")
                    st.write(f"**➖ N/A:** {row.get('items_na',0)}")
                    if row.get("observaciones"):
                        st.write(f"**📝 Obs:** {row['observaciones']}")

                # Detalle de ítems
                items_df = db.obtener_items_inspeccion(vid)
                if not items_df.empty:
                    st.markdown("##### 📋 Detalle del Checklist")
                    for sec in items_df["seccion"].unique():
                        st.markdown(f"**{sec}**")
                        sec_items = items_df[items_df["seccion"] == sec]
                        for _, it in sec_items.iterrows():
                            resp_v = it["respuesta"]
                            badge = f'<span class="badge-cumple">C</span>' if resp_v == "C" else (
                                    f'<span class="badge-nocumple">NC</span>' if resp_v == "NC" else
                                    f'<span class="badge-na">N/A</span>')
                            obs_txt = f" — <em>{it['observacion_item']}</em>" if it.get("observacion_item") else ""
                            st.markdown(
                                f'<div class="item-pregunta">{badge} {it["pregunta"]}{obs_txt}</div>',
                                unsafe_allow_html=True
                            )

                bc1, bc2 = st.columns(2)
                with bc2:
                    if st.button("🗑️ Eliminar Inspección", key=f"del_{vid}"):
                        db.eliminar_inspeccion(vid)
                        st.success("Inspección eliminada.")
                        st.rerun()
        else:
            st.warning("No hay inspecciones con los filtros seleccionados.")

    # ======================== TAB 3: DASHBOARD ========================
    with tab3:
        st.markdown("### 📊 Dashboard de Inspecciones")

        try:
            import plotly.express as px
            import plotly.graph_objects as go

            rango = st.date_input(
                "Período",
                value=(datetime.now().replace(day=1), datetime.now()),
                key="dash_rango"
            )
            if not (isinstance(rango, (list, tuple)) and len(rango) == 2):
                st.info("Selecciona un rango de fechas completo.")
                return

            df_s = db.stats_dashboard(rango[0], rango[1])

            if df_s.empty:
                st.info("No hay datos en este período.")
                return

            total = len(df_s)
            aprobadas = len(df_s[df_s["porcentaje_cumplimiento"] >= 80])
            con_obs   = len(df_s[df_s["porcentaje_cumplimiento"] < 80])
            pct_prom  = df_s["porcentaje_cumplimiento"].mean()
            total_nc  = df_s["items_no_cumple"].sum()

            k1, k2, k3, k4, k5 = st.columns(5)
            k1.metric("🔧 Total Inspecciones", total)
            k2.metric("✅ Aprobadas (≥80%)", aprobadas)
            k3.metric("❌ Con Observaciones", con_obs)
            k4.metric("📊 % Promedio", f"{pct_prom:.1f}%")
            k5.metric("⚠️ Total NC Detectados", int(total_nc))

            st.divider()

            g1, g2 = st.columns(2)
            with g1:
                st.markdown("#### Cumplimiento por Máquina")
                df_maq = df_s.groupby("maquina")["porcentaje_cumplimiento"].mean().reset_index().sort_values("porcentaje_cumplimiento")
                fig1 = px.bar(df_maq, x="porcentaje_cumplimiento", y="maquina", orientation="h",
                              color="porcentaje_cumplimiento",
                              color_continuous_scale=["#e74c3c", "#f39c12", "#2ecc71"],
                              range_color=[0, 100],
                              text=df_maq["porcentaje_cumplimiento"].apply(lambda x: f"{x:.1f}%"))
                fig1.update_traces(textposition="outside")
                fig1.update_layout(margin=dict(t=10, b=10), height=350,
                                   coloraxis_showscale=False,
                                   xaxis_title="% Cumplimiento", yaxis_title="",
                                   xaxis_range=[0, 115])
                st.plotly_chart(fig1, use_container_width=True)

            with g2:
                st.markdown("#### Inspecciones por Día")
                df_dia = df_s.groupby("fecha").size().reset_index(name="inspecciones")
                fig2 = px.bar(df_dia, x="fecha", y="inspecciones",
                              color_discrete_sequence=["#0f3460"], text="inspecciones")
                fig2.update_traces(textposition="outside")
                fig2.update_layout(margin=dict(t=10, b=10), height=350,
                                   xaxis_title="", yaxis_title="Inspecciones")
                st.plotly_chart(fig2, use_container_width=True)

            st.divider()

            g3, g4 = st.columns(2)
            with g3:
                st.markdown("#### Tendencia de Cumplimiento")
                df_trend = df_s.groupby("fecha")["porcentaje_cumplimiento"].mean().reset_index()
                fig3 = px.line(df_trend, x="fecha", y="porcentaje_cumplimiento",
                               markers=True, line_shape="spline",
                               color_discrete_sequence=["#f5a623"])
                fig3.add_hline(y=80, line_dash="dash", line_color="red",
                               annotation_text="Mín 80%", annotation_position="bottom right")
                fig3.update_layout(margin=dict(t=10, b=10), height=300,
                                   yaxis_range=[0, 105],
                                   xaxis_title="", yaxis_title="% Cumplimiento")
                st.plotly_chart(fig3, use_container_width=True)

            with g4:
                st.markdown("#### NC por Sección")
                df_s2 = db.obtener_inspecciones(rango[0], rango[1])
                nc_sec = {}
                for _, fi_row in df_s2.iterrows():
                    items_d = db.obtener_items_inspeccion(int(fi_row["id"]))
                    if not items_d.empty:
                        nc = items_d[items_d["respuesta"] == "NC"]
                        for _, r in nc.iterrows():
                            sec = str(r["seccion"])
                            nc_sec[sec] = nc_sec.get(sec, 0) + 1
                if nc_sec:
                    df_nc_sec = pd.DataFrame(list(nc_sec.items()), columns=["Sección", "NC"])
                    fig4 = px.pie(df_nc_sec, values="NC", names="Sección", hole=0.4,
                                  color_discrete_sequence=["#e74c3c", "#f39c12", "#3498db", "#2ecc71"])
                    fig4.update_layout(margin=dict(t=10, b=10), height=300)
                    st.plotly_chart(fig4, use_container_width=True)
                else:
                    st.info("Sin datos de NC en este período.")

            st.divider()

            st.markdown("#### 🏆 Ranking de Cumplimiento por Operador")
            df_op = df_s[df_s["operador"].notna() & (df_s["operador"].str.strip() != "")].groupby("operador").agg(
                inspecciones=("operador", "count"),
                pct_prom=("porcentaje_cumplimiento", "mean"),
                nc_total=("items_no_cumple", "sum"),
            ).reset_index().sort_values("pct_prom", ascending=False)
            df_op["% Cumplimiento"] = df_op["pct_prom"].apply(lambda x: f"{x:.1f}%")
            df_op = df_op.rename(columns={"operador": "Operador", "inspecciones": "Inspecciones", "nc_total": "NC Total"})
            df_op = df_op[["Operador", "Inspecciones", "% Cumplimiento", "NC Total"]]
            st.dataframe(df_op, use_container_width=True, hide_index=True)

        except ImportError:
            st.warning("Instala plotly: `pip install plotly`")
        except Exception as e:
            st.error(f"Error en dashboard: {e}")


if __name__ == "__main__":
    main()
